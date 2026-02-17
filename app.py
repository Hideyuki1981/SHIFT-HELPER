import streamlit as st
import pandas as pd
from ortools.sat.python import cp_model
import datetime
import calendar
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =====================
# マニュアルの本文
# =====================
MANUAL_TEXT = """
### 1. 基本ルール
* **ファイル名**: `staff.xlsx`
* **シート構成**: 
    1. `staff`: スタッフ設定（必須）
    2. `ng_pair`: 禁止ペア設定（任意・新規追加）

### 1. 基本ルール
* **ファイル名**: `staff.xlsx`
* **シート構成**: 
    1. `staff`: スタッフ設定（必須）
    2. `ng_pair`: 禁止ペア設定（任意・英語のまま）

### 2. staffシートの入力内容 (日本語カラム名)
| カラム名 | 説明 |
| :--- | :--- |
| **名前** | スタッフの名前 (必須) |
| **属性** | `常勤` または `派遣` |
| **公休数** | 月に必要な公休の日数 |
| **早 / 日 / 遅 / 夜** | 各シフトの可否 (`1`=可, `0`=不可) |
| **月 ～ 日** | 曜日ごとの可否 (`1`=可, `0`=不可) |
| **夜勤専従** | `1`=夜勤専従 (優先的に夜勤を割当) |
| **最大連勤** | 最大連勤数 (空欄=6連勤まで) |
| **前月夜** | 前月末が夜勤なら`1` |
| **前月連勤** | 前月末時点の連勤数 |
| **夜勤上限** | 夜勤の最大回数 (空欄=制限なし) |
| **1 ～ 31** | 日付ごとの希望 (`休`,`有`,`早`など) |

### 3. ng_pairシート（禁止ペア）
相性の悪いペアなどを入力します。可能な限り同日勤（早/日/遅）を避けます。
※「早番」と「夜勤」のような時間帯が被らない組み合わせはOKとみなされます。

| 列名 | 説明 |
| :--- | :--- |
| **staff1** | スタッフAの名前 |
| **staff2** | スタッフBの名前 |

### 4. 希望シフト
* **req_shift_1 ～ 31** に入力します
* **空欄**: おまかせ
* **休**: 希望休 (黄色)
* **有**: 有給 (オレンジ)
* **早/日/遅/夜**: シフト固定
"""

# =====================
# ページ設定
# =====================
st.set_page_config(page_title="シフト自動作成ツール", layout="wide")

st.title("📅 シフト自動作成ツール (Ver.9.3)")
st.markdown("スタッフ設定ファイル(Excel)をアップロードして、作成ボタンを押してください。")

# =====================
# サイドバー
# =====================
with st.sidebar:
    st.header("設定")
    with st.expander("📖 使い方マニュアルを見る"):
        st.markdown(MANUAL_TEXT)
    st.markdown("---")
    
    default_year = datetime.date.today().year
    default_month = datetime.date.today().month + 1
    if default_month > 12:
        default_year += 1
        default_month = 1
        
    YEAR = st.number_input("作成する【年】", value=default_year, step=1)
    MONTH = st.number_input("作成する【月】", value=default_month, min_value=1, max_value=12, step=1)
    
    uploaded_file = st.file_uploader("staff.xlsx をアップロード", type=["xlsx"])
    
    st.markdown("---")
    st.write("👥 **曜日別の必要人数設定**")
    st.caption("セルをダブルクリックして値を変更できます")

    # デフォルト値の設定 (行:シフト, 列:曜日)
    default_req_data = {
        "月": [1, 1, 1, 2],
        "火": [1, 1, 1, 2],
        "水": [1, 1, 1, 2],
        "木": [1, 1, 1, 2],
        "金": [1, 1, 1, 2],
        "土": [1, 1, 1, 2],
        "日": [1, 1, 1, 2]
    }
    df_req_default = pd.DataFrame(default_req_data, index=["早", "日", "遅", "夜"])
    
    # データエディタで編集可能にする
    edited_req_df = st.data_editor(df_req_default, height=180)

    st.markdown("---")
    st.write("🔧 **詳細設定**")
    random_seed = st.number_input("再計算用乱数 (結果を変えたい時は数字を変更)", value=0, step=1)

# =====================
# ヘルパー関数: モデル作成
# =====================
def create_shift_model(staff_df, ng_pairs, year, month, req_df, is_diagnostic=False):
    """
    モデルを作成する関数
    req_df: 曜日ごとの必要人数が入ったDataFrame
    ng_pairs: 禁止ペアのリスト [(Aさん, Bさん), ...]
    """
    DAYS = calendar.monthrange(int(year), int(month))[1]
    SHIFT_TYPES = ["早", "日", "遅", "夜"]
    # 日勤帯として扱うシフト（常勤確保、NGペア判定用）
    DAY_SHIFT_GROUP = ["早", "日", "遅"] 
    WEEKDAY_NAMES = ["月", "火", "水", "木", "金", "土", "日"]
    
    base_date = datetime.date(int(year), int(month), 1)
    weekdays_indices = [(base_date + datetime.timedelta(days=i)).weekday() for i in range(DAYS)]

    staff = staff_df["name"].tolist()
    staff_type = dict(zip(staff_df["name"], staff_df["type"]))
    off_days = dict(zip(staff_df["name"], staff_df["off_days"]))

    # 希望入力の整理
    req_input = {s: {} for s in staff}
    for _, r in staff_df.iterrows():
        s = r["name"]
        for d in range(DAYS):
            val = r.get(f"req_shift_{d+1}", None)
            if pd.isna(val) or val == "" or val == 0:
                req_input[s][d] = None
            else:
                req_input[s][d] = str(val).strip()

    prev_night = dict(zip(staff_df["name"], staff_df["prev_night"] if "prev_night" in staff_df.columns else [0]*len(staff)))
    prev_consecutive = dict(zip(staff_df["name"], staff_df["prev_consecutive"] if "prev_consecutive" in staff_df.columns else [0]*len(staff)))
    night_only = dict(zip(staff_df["name"], staff_df["night_only"] if "night_only" in staff_df.columns else [0]*len(staff)))

    can = {}
    weekday_can = {}
    for _, r in staff_df.iterrows():
        s = r["name"]
        can[s] = {"早": r["can_early"], "日": r["can_day"], "遅": r["can_late"], "夜": r["can_night"]}
        weekday_can[s] = {0: r["mon"], 1: r["tue"], 2: r["wed"], 3: r["thu"], 4: r["fri"], 5: r["sat"], 6: r["sun"]}

    limit_consecutive = dict(zip(staff_df["name"], staff_df["limit_consecutive"] if "limit_consecutive" in staff_df.columns else [6]*len(staff)))

    # --- モデル構築開始 ---
    model = cp_model.CpModel()
    x = {}
    work_flag = {}

    # 変数定義
    for s in staff:
        for d in range(DAYS):
            for sh in SHIFT_TYPES:
                if can[s][sh] == 1:
                    x[s, d, sh] = model.NewBoolVar(f"x_{s}_{d}_{sh}")
                else:
                    x[s, d, sh] = None 

    # 1日1シフトまで
    for s in staff:
        for d in range(DAYS):
            model.Add(sum(x[s, d, sh] for sh in SHIFT_TYPES if x[s, d, sh] is not None) <= 1)

    # ==========================================
    # 人数制約 (曜日別の設定値を参照)
    # ==========================================
    shortage_vars = {} 

    for d in range(DAYS):
        wd_idx = weekdays_indices[d]     # 0~6
        wd_name = WEEKDAY_NAMES[wd_idx]  # "月"~"日"
        
        for sh in SHIFT_TYPES:
            active_staff = sum(x[s, d, sh] for s in staff if x[s, d, sh] is not None)
            
            # DataFrameからその曜日・そのシフトの必要人数を取得
            req_num = int(req_df.at[sh, wd_name])

            if not is_diagnostic:
                model.Add(active_staff == req_num)
            else:
                shortage = model.NewIntVar(0, req_num, f"shortage_{d}_{sh}")
                shortage_vars[d, sh] = shortage
                model.Add(active_staff + shortage == req_num)

    # 勤務フラグ作成
    for s in staff:
        for d in range(DAYS):
            work_flag[s, d] = model.NewBoolVar(f"work_{s}_{d}")
            is_shifted = sum(x[s, d, sh] for sh in SHIFT_TYPES if x[s, d, sh] is not None)
            if d == 0:
                is_after_night = prev_night[s]
            else:
                is_after_night = x[s, d-1, "夜"] if x[s, d-1, "夜"] is not None else 0
            model.Add(work_flag[s, d] == is_shifted + is_after_night)

    # 希望反映
    for s in staff:
        for d in range(DAYS):
            inp = req_input[s][d]
            if inp is None: continue
            if inp == "休" or inp == "有":
                model.Add(sum(x[s, d, sh] for sh in SHIFT_TYPES if x[s, d, sh] is not None) == 0)
                model.Add(work_flag[s, d] == 0)
            elif inp in SHIFT_TYPES:
                if x[s, d, inp] is not None:
                    model.Add(x[s, d, inp] == 1)
                    for other_sh in SHIFT_TYPES:
                        if other_sh != inp and x[s, d, other_sh] is not None:
                            model.Add(x[s, d, other_sh] == 0)
                else:
                    model.Add(work_flag[s, d] == 0) 

    # 夜勤ルール
    next_day_off_penalty = [] 
    for s in staff:
        for d in range(DAYS):
            if x[s, d, "夜"] is not None:
                # 既存: 翌日(d+1)は明け（勤務なし）
                if d + 1 < DAYS:
                    model.Add(sum(x[s, d+1, sh] for sh in SHIFT_TYPES if x[s, d+1, sh] is not None) == 0).OnlyEnforceIf(x[s, d, "夜"])
                
                # ==========================
                # 【修正箇所】夜明の翌日(d+2)に早番を禁止
                # ==========================
                if d + 2 < DAYS:
                    # そのスタッフが早番可能(Noneでない)であれば、夜勤の翌々日の早番を0に固定
                    if x[s, d+2, "早"] is not None:
                        model.Add(x[s, d+2, "早"] == 0).OnlyEnforceIf(x[s, d, "夜"])

                # 既存: 翌々日の勤務間隔制約（変更なし）
                if d + 2 < DAYS:
                    violation = model.NewBoolVar(f"violation_{s}_{d}")
                    model.AddBoolAnd([x[s, d, "夜"], work_flag[s, d+2]]).OnlyEnforceIf(violation)
                    model.AddBoolOr([x[s, d, "夜"].Not(), work_flag[s, d+2].Not()]).OnlyEnforceIf(violation.Not())
                    next_day_off_penalty.append(violation)

    # ==========================================
    # 【修正】連休制限: 「常勤」のみ、自動割り当ての「休」は連続2回まで
    # ==========================================
    for s in staff:
        # 派遣や非常勤は休みが多くてもOKなのでスキップ
        if staff_type.get(s, "") != "常勤":
            continue

        for d in range(DAYS - 2):
            # 連続する3日間の希望内容を取得
            check_reqs = [req_input[s][d], req_input[s][d+1], req_input[s][d+2]]
            
            # 「休(希望休)」または「有(有給)」が1つでも含まれていれば、この期間の連休制限はしない
            # (ユーザーが意図して連休を入れた場合は許可)
            if any(r in ["休", "有"] for r in check_reqs):
                continue
            
            # 3日間すべて work_flag=0 (勤務なし) になることを禁止
            # これにより、勤務(1)が含まれない「休・休・休」の並びを防ぐ
            model.Add(sum(work_flag[s, d+k] for k in range(3)) >= 1)

       # 曜日制限
    for s in staff:
        for d in range(DAYS):
            if weekday_can[s][weekdays_indices[d]] == 0:
                model.Add(sum(x[s, d, sh] for sh in SHIFT_TYPES if x[s, d, sh] is not None) == 0)

    # 公休数
    for s in staff:
        total_work = sum(work_flag[s, d] for d in range(DAYS))
        paid_leave = sum(1 for d in range(DAYS) if req_input[s][d] == "有")
        if staff_type[s] == "常勤":
            model.Add(total_work == DAYS - off_days[s] - paid_leave)
        else:
            model.Add(total_work <= DAYS - off_days[s] - paid_leave)

    # ==========================================
    # 追加機能1: 常勤スタッフ確保 (早・日・遅に最低1名)
    # ==========================================
    full_time_staff = [s for s in staff if staff_type.get(s, "") == "常勤"]
    no_leader_penalty = []

    if full_time_staff:
        for d in range(DAYS):
            # その日の日勤帯(早日遅)に入っている常勤スタッフの数
            # x変数がNone(シフト不可)でない場合のみ合計する
            ft_count_vars = []
            for s in full_time_staff:
                for sh in DAY_SHIFT_GROUP:
                    if x[s, d, sh] is not None:
                        ft_count_vars.append(x[s, d, sh])
            
            if ft_count_vars:
                # 常勤が0人になったらフラグを立てる
                is_no_leader = model.NewBoolVar(f"no_leader_{d}")
                model.Add(sum(ft_count_vars) == 0).OnlyEnforceIf(is_no_leader)
                model.Add(sum(ft_count_vars) >= 1).OnlyEnforceIf(is_no_leader.Not())
                no_leader_penalty.append(is_no_leader)

    # ==========================================
    # 追加機能2: NGペア制約 (日勤帯のかぶり禁止)
    # ==========================================
    ng_pair_penalty = []
    
    for (p1, p2) in ng_pairs:
        # 両方のスタッフが存在する場合のみチェック
        if p1 in staff and p2 in staff:
            for d in range(DAYS):
                # p1が日勤帯(早日遅)にいるか
                p1_vars = [x[p1, d, sh] for sh in DAY_SHIFT_GROUP if x[p1, d, sh] is not None]
                p1_is_day = model.NewBoolVar(f"ng_{p1}_{d}")
                if p1_vars:
                    model.Add(sum(p1_vars) == 1).OnlyEnforceIf(p1_is_day)
                    model.Add(sum(p1_vars) == 0).OnlyEnforceIf(p1_is_day.Not())
                else:
                    model.Add(p1_is_day == 0)

                # p2が日勤帯(早日遅)にいるか
                p2_vars = [x[p2, d, sh] for sh in DAY_SHIFT_GROUP if x[p2, d, sh] is not None]
                p2_is_day = model.NewBoolVar(f"ng_{p2}_{d}")
                if p2_vars:
                    model.Add(sum(p2_vars) == 1).OnlyEnforceIf(p2_is_day)
                    model.Add(sum(p2_vars) == 0).OnlyEnforceIf(p2_is_day.Not())
                else:
                    model.Add(p2_is_day == 0)

                # 両方日勤帯ならNG
                is_ng_clash = model.NewBoolVar(f"ng_clash_{p1}_{p2}_{d}")
                model.AddBoolAnd([p1_is_day, p2_is_day]).OnlyEnforceIf(is_ng_clash)
                model.AddBoolOr([p1_is_day.Not(), p2_is_day.Not()]).OnlyEnforceIf(is_ng_clash.Not())
                
                ng_pair_penalty.append(is_ng_clash)

    # ==========================================
    # 目的関数
    # ==========================================
    # 【追加】夜勤回数制限の読み込み (列がない場合は無視)
    staff_max_night = {}
    if "max_night" in staff_df.columns:
        staff_max_night = dict(zip(staff_df["name"], staff_df["max_night"]))
    else:
        staff_max_night = {s: None for s in staff}

    night_count = {}
    for s in staff:
        night_count[s] = model.NewIntVar(0, DAYS, f"night_{s}")
        model.Add(night_count[s] == sum(x[s, d, "夜"] for d in range(DAYS) if x[s, d, "夜"] is not None))
        
        # 【追加】個別の回数制限がある場合、その回数以下にする
        limit = staff_max_night.get(s)
        if pd.notna(limit) and limit != "":
            model.Add(night_count[s] <= int(limit))
    
    night_penalty = []
    dispatch_penalty = []
    night_maximization_bonus = []
    balance_penalty = []

    target_workers = [s for s in staff if can[s]["夜"] == 1 and night_only.get(s, 0) == 0]
    if target_workers:
        max_night = model.NewIntVar(0, DAYS, "max_night")
        min_night = model.NewIntVar(0, DAYS, "min_night")
        for s in target_workers:
            model.Add(night_count[s] <= max_night)
            model.Add(night_count[s] >= min_night)
        diff = model.NewIntVar(0, DAYS, "night_diff")
        model.Add(diff == max_night - min_night)
        balance_penalty.append(diff)

    for s in staff:
        for d in range(DAYS):
            if x[s, d, "夜"] is not None and night_only.get(s, 0) == 0:
                night_penalty.append(x[s, d, "夜"])
            if staff_type[s] == "派遣":
                for sh in SHIFT_TYPES:
                    if x[s, d, sh] is not None:
                        dispatch_penalty.append(x[s, d, sh])
    
    night_only_staff = [s for s in staff if night_only.get(s, 0) == 1]
    for s in night_only_staff:
        night_maximization_bonus.append(night_count[s])

    if not is_diagnostic:
        # 重みづけ設定
        # no_leader_penalty: 常勤不在はかなり避けたい (200)
        # ng_pair_penalty: NGペアも避けたいが、人員不足よりはマシ (100)
        model.Minimize(
            10 * sum(night_penalty) +
            100 * sum(dispatch_penalty) +
            100 * sum(balance_penalty) +
            500 * sum(next_day_off_penalty) + 
            200 * sum(no_leader_penalty) + 
            100 * sum(ng_pair_penalty)
            - 1000 * sum(night_maximization_bonus)
        )
    else:
        total_shortage_val = sum(shortage_vars.values())
        model.Minimize(total_shortage_val)

    return model, x, shortage_vars, req_input, prev_night, staff

# =====================
# メイン処理
# =====================
if st.button("シフトを作成する", type="primary"):
    if uploaded_file is None:
        st.error("エラー: Excelファイルをアップロードしてください。")
        st.stop()

    try:
        DAYS = calendar.monthrange(int(YEAR), int(MONTH))[1]
        st.info(f"{YEAR}年{MONTH}月 ({DAYS}日分) のシフトを計算中...")

        SHIFT_TYPES = ["早", "日", "遅", "夜"]

        # Excel読み込み
        staff_df = pd.read_excel(uploaded_file, sheet_name="staff")
        
        # ==========================================
        # 【追加】日本語カラム名を内部用(英語)に変換
        # ==========================================
        jp_to_en = {
            "名前": "name",
            "属性": "type",
            "公休数": "off_days",
            "早番": "can_early",
            "日勤": "can_day",
            "遅番": "can_late",
            "夜勤": "can_night",
            "月": "mon", "火": "tue", "水": "wed", "木": "thu", "金": "fri", "土": "sat", "日": "sun",
            "夜勤専従": "night_only",
            "最大連勤": "limit_consecutive",
            "前月末夜": "prev_night",
            "前月連勤": "prev_consecutive",
            "夜勤上限": "max_night"
        }
        staff_df = staff_df.rename(columns=jp_to_en)

        # 日付カラム (1, 2, ... 31) を req_shift_1, req_shift_2 ... に変換
        # ※Excelで数字の「1」と入力されても、文字の「1」と入力されても対応
        for d in range(1, 32):
            if d in staff_df.columns:
                staff_df = staff_df.rename(columns={d: f"req_shift_{d}"})
            elif str(d) in staff_df.columns:
                staff_df = staff_df.rename(columns={str(d): f"req_shift_{d}"})
        
        # NGペアシートの読み込み (存在しない場合のエラーハンドリング)
        ng_pairs = []
        try:
            ng_df = pd.read_excel(uploaded_file, sheet_name="ng_pair")
            # 空白行除去や列存在チェック
            if "staff1" in ng_df.columns and "staff2" in ng_df.columns:
                ng_pairs = list(zip(ng_df["staff1"], ng_df["staff2"]))
            else:
                st.warning("⚠️ `ng_pair` シートが見つかりましたが、列名 `staff1`, `staff2` が正しくありません。NGペア設定はスキップされます。")
        except ValueError:
            # シートがない場合は何もしない
            pass
        except Exception as e:
            st.warning(f"⚠️ NGペア設定の読み込み中にエラーが発生しました: {e}")

        
        # --- 1回目: 通常計算 ---
        model, x, _, req_input, prev_night, staff = create_shift_model(staff_df, ng_pairs, YEAR, MONTH, edited_req_df, is_diagnostic=False)
        
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 30.0 
        solver.parameters.random_seed = int(random_seed)

        status = solver.Solve(model)

        if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            st.success(f"✅ 作成成功！ (Status: {solver.StatusName(status)})")

            # 結果作成処理
            count_cols = ["早", "日", "遅", "夜", "休", "有"]
            count_data = {c: [] for c in count_cols}
            result = []

            for s in staff:
                row = []
                p_counts = {c: 0 for c in count_cols}
                
                for d in range(DAYS):
                    v = ""
                    if req_input[s][d] == "有":
                        v = "有"
                        p_counts["有"] += 1
                    else:
                        is_work = False
                        for sh in SHIFT_TYPES:
                            if x[s, d, sh] is not None and solver.Value(x[s, d, sh]) == 1:
                                v = sh
                                p_counts[sh] += 1
                                is_work = True
                                break
                        
                        if not is_work:
                            prev_is_night = False
                            if d > 0:
                                if x[s, d-1, "夜"] is not None and solver.Value(x[s, d-1, "夜"]) == 1:
                                    prev_is_night = True
                            elif d == 0:
                                if prev_night[s] == 1:
                                    prev_is_night = True
                            
                            if prev_is_night:
                                v = "明"
                            else:
                                v = "休"
                                p_counts["休"] += 1
                    
                    row.append(v)
                result.append(row)
                
                for c in count_cols:
                    count_data[c].append(p_counts[c])
            
            df_out = pd.DataFrame(result, index=staff, columns=[f"{i+1}日" for i in range(DAYS)])
            for c in count_cols:
                df_out[c] = count_data[c]

            def highlight_cells(val):
                # 薄い緑 (#E2F0D9)
                if val in ["明", "休", "有"]:
                    return "background-color: #E2F0D9; color: black"
                return ""

            st.dataframe(df_out.style.map(highlight_cells))

            # Excel出力
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_out.to_excel(writer, sheet_name='Result')
            
            output.seek(0)
            wb = load_workbook(output)
            ws = wb.active
            
            # 薄い緑の定義
            green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

            # ==========================================
            # 【修正】Excel色塗り: 明・休・有 をすべて緑に統一
            # ==========================================
            for i, s in enumerate(staff, start=2):
                for d in range(DAYS):
                    cell = ws.cell(row=i, column=d+2)
                    val = cell.value
                    
                    # セルの文字が「明」「休」「有」のいずれかなら緑にする
                    if val in ["明", "休", "有"]:
                        cell.fill = green_fill

            final_output = BytesIO()
            wb.save(final_output)
            final_output.seek(0)

            st.download_button(
                label="📥 Excelファイルをダウンロード",
                data=final_output,
                file_name=f"shift_{YEAR}_{MONTH}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        else:
            # --- 2回目: 診断モード ---
            st.error("❌ 条件を満たすシフトが見つかりませんでした。")
            st.warning("🕵️‍♂️ 原因を調査しています...（不足している人員箇所を計算中）")

            # 診断モードではNGペアなどは無視して、人数の絶対値不足を見るため ng_pairs=[] で渡す
            diag_model, diag_x, shortage_vars, _, _, _ = create_shift_model(staff_df, [], YEAR, MONTH, edited_req_df, is_diagnostic=True)
            
            diag_solver = cp_model.CpSolver()
            diag_status = diag_solver.Solve(diag_model)

            if diag_status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
                error_rows = []
                WEEKDAY_NAMES = ["月", "火", "水", "木", "金", "土", "日"]
                base_date = datetime.date(int(YEAR), int(MONTH), 1)

                for d in range(DAYS):
                    date_obj = base_date + datetime.timedelta(days=d)
                    wd_idx = date_obj.weekday()
                    wd_str = WEEKDAY_NAMES[wd_idx]

                    for sh in SHIFT_TYPES:
                        lack = diag_solver.Value(shortage_vars[d, sh])
                        if lack > 0:
                            req_val = int(edited_req_df.at[sh, wd_str])
                            actual_val = req_val - lack
                            error_rows.append({
                                "date": f"{d+1}日({wd_str})",
                                "shift": sh,
                                "req": req_val,
                                "act": actual_val,
                                "lack": lack
                            })
                
                # ■■■ 変更点：具体的な箇条書き表示 ■■■
                if error_rows:
                    st.markdown("### ⚠️ 人員不足レポート")
                    st.error(f"合計 {len(error_rows)} 箇所で人員が足りていません。")
                    
                    for row in error_rows:
                        # 読みやすい形式で出力
                        st.markdown(
                            f"- **{row['date']}** : **【{row['shift']}】** が **{row['lack']}名** 不足 "
                            f"(必要: {row['req']}名 → 確保可能: {row['act']}名)"
                        )
                    
                    st.markdown("---")
                    st.info("💡 **対策**: 上記の日付の「必要人数」を減らすか、スタッフの「希望休」を取り下げて調整してください。")
                else:
                    st.warning("人員数は足りていますが、その他の制約（連勤制限やNGペアなど）で矛盾が生じています。")
            else:
                st.error("人員数を無視してもシフトが組めません。Excelの入力ミスがないか確認してください。")

    except Exception as e:
        st.error(f"システムエラーが発生しました: {e}")



